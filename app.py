from flask import Flask, render_template, request, g, send_file, url_for
import linearRegressionML as lr
from DatasetCardiovascular import SaludModel
from ExcelProcessor import ExcelProcessor
import sqlite3
import traceback  
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenPyXLImage
from sklearn.linear_model import LogisticRegression
from werkzeug.utils import secure_filename
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score, confusion_matrix, classification_report
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
from io import BytesIO
import seaborn as sns
from datetime import datetime
import tempfile

app = Flask(__name__)
modelo = SaludModel()
UPLOAD_FOLDER = 'archivos_cargados'
PROCESSED_FOLDER = 'static/archivos_procesados'
ALLOWED_EXTENSIONS = {'xlsx', 'csv'}
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)
DATABASE = 'modelos.db'

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route("/")
def home():
    return render_template("index.html")

@app.route("/diferencias-ia-ml")
def IaMlDifferences():
    return render_template("diferencias-ia-ml.html")

@app.route('/casos-de-exito-ml')
def successCases():
    casesList = [
        {
            'title': 'Diagnóstico Médico Asistido por IA',
            'sector': 'Salud',
            'problem': 'Desafíos en la detección temprana y precisa de enfermedades complejas.',
            'algorithm': 'Redes Neuronales Artificiales (ANN)',
            'benefits': 'Mejora en la precisión diagnóstica y detección temprana de enfermedades.',
            'company': 'IBM Watson analiza datos médicos para asistir en diagnósticos más precisos y rápidos.',
            'reference': 'Imagar. (2024). Machine Learning: casos de éxito en diferentes sectores. Recuperado de https://www.imagar.com/informatica/machine-learning-casos-de-exito-en-diferentes-sectores/',
            'referenceUrl': 'https://www.imagar.com/informatica/machine-learning-casos-de-exito-en-diferentes-sectores/'
        },
        {
            'title': 'Sistemas de Recomendación en Plataformas de Streaming',
            'sector': 'Entretenimiento',
            'problem': 'Necesidad de personalizar la experiencia del usuario mediante recomendaciones de contenido.',
            'algorithm': 'Sistemas de Recomendación basados en Filtrado Colaborativo',
            'benefits': 'Incremento en el tiempo de visualización y lealtad del cliente.',
            'company': 'Netflix utiliza algoritmos de Machine Learning para ofrecer recomendaciones basadas en el comportamiento del usuario.',
            'reference': 'Beservices. (2020). Ejemplos de Machine Learning en empresas. Recuperado de https://blog.beservices.es/blog/ejemplos-de-machine-learning-en-empresas',
            'referenceUrl': 'https://blog.beservices.es/blog/ejemplos-de-machine-learning-en-empresas'
        },
        {
            'title': 'Optimización de Rutas en Servicios de Logística',
            'sector': 'Logística y Transporte',
            'problem': 'Minimizar tiempos de entrega y costos operativos en rutas de transporte.',
            'algorithm': 'Algoritmos de Optimización y Aprendizaje Supervisado',
            'benefits': 'Reducción del consumo de combustible y mejora en los tiempos de entrega.',
            'company': 'UPS utiliza Machine Learning para programar rutas que minimizan los giros a la izquierda, optimizando el tiempo de reparto.',
            'reference': 'Beservices. (2020). Ejemplos de Machine Learning en empresas. Recuperado de https://blog.beservices.es/blog/ejemplos-de-machine-learning-en-empresas',
            'referenceUrl': 'https://blog.beservices.es/blog/ejemplos-de-machine-learning-en-empresas'
        }
    ]
    return render_template('casos-de-exito-ml.html', cases=casesList)

@app.route("/regresion-lineal", methods=["GET", "POST"])
def calculateGrade():
    calculateResult = None
    plot_url = None
    hours = None

    if request.method == "POST":
        try:
            hours = float(request.form["Hours"])
            calculateResult = lr.calculateGrade(hours)
            plot_url = lr.generate_plot(hours)  # Generar gráfica solo si el input es válido
        except ValueError:
            calculateResult = "Invalid input. Please enter a number."
            plot_url = None  # Evita mostrar una gráfica si hay error en la entrada
    
    return render_template("regresion-lineal-prediccion-notas.html", result=calculateResult, plot_url=plot_url, hours=hours)

@app.route("/mapa-mental")
def mindMap():
    return render_template("mapa-mental-regresion-logistica.html")

@app.route('/regresion-logistica')  
def regresion_logistica():
    try:
        # Inicializar modelo
        modelo = SaludModel()  
        
        # 1. Generar y preparar datos
        df = modelo.generar_datos()
        if df.empty:
            raise ValueError("El DataFrame generado está vacío")
        
        # 2. Entrenar modelo y obtener métricas
        metricas = modelo.entrenar_modelo()
        required_metrics = ['accuracy', 'precision', 'recall', 'f1', 'roc_auc', 'confusion', 'y_pred']
        if not all(m in metricas for m in required_metrics):
            missing = [m for m in required_metrics if m not in metricas]
            raise ValueError(f"Métricas faltantes: {', '.join(missing)}")

        # 3. Generar gráficos (con manejo de errores individual)
        graficos = {}
        try:
            graficos['edad_presion'] = modelo.generar_grafico_edad_presion()
            graficos['confusion'] = modelo.generar_grafico_confusion()
            graficos['roc'] = modelo.generar_curva_roc()
            graficos['metricas'] = modelo.generar_grafico_metricas()
            graficos['coeficientes'] = modelo.generar_grafico_coeficientes()
        except Exception as e:
            print(f"Error generando gráficos: {str(e)}")
            # Puedes continuar aunque falle algún gráfico

        # 4. Preparar datos para la tabla
        tabla_datos = modelo.get_datos_para_tabla()
        if not tabla_datos:
            print("Advertencia: Tabla de datos vacía")
            tabla_datos = []

        # 5. Calcular estadísticas
        stats = {
            "total": len(df),
            "enfermos": int(df["Enfermedad"].sum()),
            "sanos": len(df) - int(df["Enfermedad"].sum()),
            "edad_promedio": round(df["Edad"].mean(), 1),
            "presion_promedio": round(df["Presion"].mean(), 1),
            "colesterol_promedio": round(df["Colesterol"].mean(), 1),
            "accuracy": round(metricas["accuracy"] * 100, 1),
            "auc": round(metricas["roc_auc"] * 100, 1)
        }

        # 6. Renderizar plantilla
        return render_template(
            'regresion-logistica-ml.html',  # Asegúrate que coincida con tu archivo
            stats=stats,
            metrics={
                "accuracy": metricas["accuracy"],
                "precision": metricas["precision"],
                "recall": metricas["recall"],
                "f1": metricas["f1"],
                "roc_auc": metricas["roc_auc"],
                "confusion": metricas["confusion"]
            },
            grafico_edad_presion=graficos.get('edad_presion'),
            grafico_confusion=graficos.get('confusion'),
            graficos={
                "roc": graficos.get('roc'),
                "coeficientes": graficos.get('coeficientes'),
                "metricas": graficos.get('metricas')
            },
            tabla_datos=tabla_datos,
            predicciones=metricas.get("y_pred", [])
        )
        
    except Exception as e:
        # Obtener información del error
        error_msg = f"Error al procesar la regresión logística: {str(e)}"
        error_details = traceback.format_exc() if app.debug else None
        
        # Registrar el error
        app.logger.error(f"Error en regresion_logistica: {error_msg}\n{error_details}")
        
        # Verificar si la plantilla existe antes de renderizar
        template_path = os.path.join(app.root_path, 'templates', 'error.html')
        if not os.path.exists(template_path):
            return f"""
            <h1>Error crítico</h1>
            <p>{error_msg}</p>
            <pre>{error_details if app.debug else 'Habilita el modo debug para ver detalles'}</pre>
            """, 500
            
        return render_template(
            'error.html',
            error_message=error_msg,
            error_details=error_details
        ), 500

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DATABASE)
    return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

@app.route('/modelo-supervisado/<int:modelo_id>')
def mostrar_modelo(modelo_id):
    cur = get_db().cursor()
    cur.execute("SELECT * FROM modelos WHERE id = ?", (modelo_id,))
    modelo = cur.fetchone()
    return render_template('modelos-supervisados.html', modelo=modelo)

@app.route('/regresion-logistica-excel', methods=['GET', 'POST'])
def subir_archivo():
    if request.method == 'POST':
        file = request.files['file']
        if file:
            filename = file.filename
            ext = os.path.splitext(filename)[1].lower()

            if ext == '.csv':
                df = pd.read_csv(file, sep=';')
            elif ext in ['.xlsx', '.xls']:
                df = pd.read_excel(file)
            else:
                return "Formato de archivo no soportado", 400

            columnas = df.columns.tolist()
            df.to_csv('archivo_temporal.csv', index=False)

            return render_template('regresion-logistica-excel/seleccionar-columna.html', columnas=columnas)

    return render_template('regresion-logistica-excel/cargar-archivo.html') 

@app.route('/ejecutar_modelo', methods=['POST'])
def ejecutar_modelo():
    target = request.form['target']
    df = pd.read_csv('archivo_temporal.csv')

    X = df.drop(columns=[target])
    y = df[target]

    model = LogisticRegression(max_iter=1000, multi_class='ovr')
    model.fit(X, y)

    y_pred = model.predict(X)
    y_proba = model.predict_proba(X)[:, 1]

    acc = accuracy_score(y, y_pred)
    prec = precision_score(y, y_pred, average='weighted')
    rec = recall_score(y, y_pred, average='weighted')
    f1 = f1_score(y, y_pred, average='weighted')
    # Matriz de confusión multiclase
    conf_matrix = confusion_matrix(y, y_pred)

    print(f"Forma de la matriz de confusión: {conf_matrix.shape}")
    print(f"Clases en y: {y.unique()}")
    print(f"Clases en y_pred: {np.unique(y_pred)}")


    df_resultado = X.copy()
    df_resultado['Columna de Análisis'] = target
    df_resultado['Valor de origen'] = y
    df_resultado['Predicción'] = y_pred
    df_resultado['Probabilidad'] = y_proba

    df_metricas = pd.DataFrame({
        'Métrica': ['Accuracy (Precisión Global)', 'Precision (Precisión en Casos Positivos)', 'Recall (Sensibilidad o Tasa de Detección)', 'F1 Score (Equilibrio entre Precisión y Detección)'],
        'Valor': [acc, prec, rec, f1]
    })

    df_confusion = pd.DataFrame(
        conf_matrix,
        index=[f'Real {c}' for c in np.unique(y)],
        columns=[f'Pred {c}' for c in np.unique(y)]
    )
    
    ts = int(datetime.now().timestamp())
    
    # Guardar el Excel en static/resultados
    excel_name = f'reporte_modelo_{ts}.xlsx'
    excel_path = os.path.join(PROCESSED_FOLDER, excel_name)

    # == Guardar en Excel == #
    output_excel = 'reporte_modelo_completo.xlsx'
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df_resultado.to_excel(writer, index=False, sheet_name='Resultados')
        df_metricas.to_excel(writer, index=False, sheet_name='Métricas')
        df_confusion.to_excel(writer, sheet_name='Matriz de Confusión')
    
    # --- Generar imágenes en archivos temporales --- #
    # Heatmap
    heatmap_path = os.path.join(PROCESSED_FOLDER, f'heatmap_{ts}.png')
    plt.figure(figsize=(8,6))
    sns.heatmap(conf_matrix, annot=True, fmt='d', cmap='Blues',
                xticklabels=np.unique(y), yticklabels=np.unique(y))
    plt.title('Matriz de Confusión')
    plt.tight_layout()
    plt.savefig(heatmap_path)
    plt.close()

    # Scatter de probabilidades
    scatter_path = os.path.join(PROCESSED_FOLDER, f'scatter_{ts}.png')
    plt.figure(figsize=(6,4))
    plt.scatter(range(len(y_proba)), y_proba, c=y, cmap='bwr', alpha=0.7)
    plt.title('Probabilidades de predicción')
    plt.ylabel('Probabilidad de clase positiva')
    plt.xlabel('Índice de muestra')
    plt.tight_layout()
    plt.savefig(scatter_path)
    plt.close()

    # Insertar heatmap en el Excel
    wb = load_workbook(excel_path)
    ws = wb['Matriz de Confusión']
    img = OpenPyXLImage(heatmap_path)
    img.anchor = 'E2'
    ws.add_image(img)
    wb.save(excel_path)
    
    # Insertar Scatter de probabilidades en el Excel
    wb = load_workbook(excel_path)
    ws = wb['Métricas']
    img = OpenPyXLImage(scatter_path)
    img.anchor = 'E2'
    ws.add_image(img)
    wb.save(excel_path)
    
    # Renderizar plantilla de resultados
    return render_template('regresion-logistica-excel/resultados-regresion.html',
                       score=f'{acc * 100:.2f}%',
                       image_url=url_for('static', filename=f'archivos_procesados/{os.path.basename(heatmap_path)}'),
                       excel_url=url_for('static', filename=f'archivos_procesados/{excel_name}'))

    # return send_file(output_excel, as_attachment=True)

def regresion_logistica_excel():
    if request.method == 'POST':
        file = request.files['file']
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            ext = filename.rsplit('.', 1)[1].lower()
            path = os.path.join(UPLOAD_FOLDER, filename)
            file.save(path)

            processor = ExcelProcessor(path, file_type=ext)
            df, metrics_df, conf_df, image_stream = processor.train_and_predict()

            output_path = os.path.join(PROCESSED_FOLDER, 'resultado_' + filename)
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Resultados')
                metrics_df.to_excel(writer, sheet_name='Métricas')
                conf_df.to_excel(writer, sheet_name='Matriz de Confusión')

            wb = load_workbook(output_path)
            ws = wb['Matriz de Confusión']
            img = OpenPyXLImage(image_stream)
            img.anchor = 'E2'
            ws.add_image(img)
            wb.save(output_path)

            return send_file(output_path, as_attachment=True)

    return render_template('regresion-logistica-excel.html')

if __name__ == "__main__":
    app.run(debug=True)
